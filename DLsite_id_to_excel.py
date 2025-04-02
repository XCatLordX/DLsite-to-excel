import asyncio
import pandas as pd
from dlsite_async import DlsiteAPI
import requests
from io import BytesIO
from openpyxl import Workbook  # 确保导入 Workbook
from openpyxl.drawing.image import Image
import sys

async def f(a):
    async with DlsiteAPI() as api:
        return await api.get_work(a)
        
data_list = []  # 假设有多个数据对象

print("一次性輸入作品id（用空格分隔）請輸入：1")
print("分次輸入作品id請輸入：2")
Mode = input()

if Mode == '1':
    print("請一次性輸入作品id（用空格分隔）： ")
    data_id = input().split()
    for i in range(0,len(data_id)):
        raw_data = asyncio.run(f(data_id[i]))
        data_list.append(raw_data)
    
elif Mode == '2':
    while True:
        data_id = str(input("(輸入空字串即結束）請輸入作品id: "))
        if data_id == "":
            break
        else:
            raw_data = asyncio.run(f(data_id))
            data_list.append(raw_data)

else:
    print("無效輸入")
    sys.exit(0) #end program

#print(data_list)


# 创建 Excel 文件
wb = Workbook()

# 处理每个数据对象
for data_idx, raw_data in enumerate(data_list, start=1):
    # 创建新工作表
    ws = wb.create_sheet(title=raw_data.work_name)  # 新工作表，命名为 Data_1, Data_2, 等
    
    # 将 Work 对象转换成字典
    data_dict = vars(raw_data)  # 使用 vars() 将 Work 对象转换为字典

    # 处理数据（转换 None、列表等）
    for key, value in data_dict.items():
        if isinstance(value, list):
            data_dict[key] = ", ".join(map(str, value))  # 列表转字符串
        elif value is None:
            data_dict[key] = ""  # None 变成空字符串

    # 将字段名称和字段值写入 Excel（第一列：字段名，第二列：字段值）
    for row_idx, (key, value) in enumerate(data_dict.items(), start=1):
        ws.cell(row=row_idx, column=1, value=key)  # 字段名称
        if key != "work_image":  # 不是图片字段，正常写入
            ws.cell(row=row_idx, column=2, value=value)

    # 处理图片字段（下载并插入）
    image_url = data_dict.get("work_image")
    if image_url:
        image_url = "https:" + image_url  # 确保 URL 正确（你的 URL 可能缺少协议）

        try:
            # 下载图片
            response = requests.get(image_url)
            if response.status_code == 200:
                img = Image(BytesIO(response.content))

                # 获取图片应放置的行列位置（找出 "work_image" 的行号）
                img_row = list(data_dict.keys()).index('work_image') + 1
                img_cell = f"B{img_row}"  # 将图片放在第二列对应行

                # 调整单元格大小以适应图片
                column_letter = 'B'
                ws.column_dimensions[column_letter].width = 20  # 调整列宽
                ws.row_dimensions[img_row].height = 80  # 调整行高

                # 插入图片到单元格
                ws.add_image(img, img_cell)

        except Exception as e:
            print("图片下载失败:", e)

# 删除默认创建的空白工作表
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 保存 Excel 文件
wb.save("result.xlsx")
print("Excel 文件已成功生成！")
